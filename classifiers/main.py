import copy
import os

import numpy as np
import openpyxl

import Band
import Image
import Label
import Mask
import RandomForest
import TimeSeries
import Utility
import XGBoost
from ClassidierValidator import ClassifierValidator


def set_empty_bands():
    for timestamp in filePaths.keys():
        for band in filePaths[timestamp].keys():
            if not os.path.isfile(filePaths[timestamp][band]):  # controlla se non è presente il file della banda
                filePaths[timestamp][band] = emptyBandPath  # sostituisce il path della banda
                # con il path della banda vuota


# leggiamo il file di configurazione
with open('configuration.txt', mode='r') as File:
    timeStamps = File.readline().split()  # legge i timestamp da file
    bands = File.readline().strip().split()  # legge le bande da file
    labelPath = File.readline().strip()  # legge da file il path del file contenente la label
    emptyBandPath = File.readline().strip()  # legge da file il path del file contenente la banda vuota
    pathVal = File.readline().strip()  # legge il path
    percentualSplit = float(File.readline().strip())  # percentuale divisione train e test set
    filePaths = {}  # dizionario con: key = timestamp, value = BandsPath relativo a quel timestamp
    bandsPath = {}  # dizionario contenente i path di ciascuna banda
    i = 0
    for image in timeStamps:
        for band in bands:
            temp = pathVal + image + "\\" + band + "_" + image + ".tif"
            bandsPath[band] = temp
        filePaths[
            image] = bandsPath  # salva il dizionario con tutti i path delle bande nel timestamp corrispondente
        bandsPath = {}

    indexes = []  # lista degli indici da calcolare
    for row in File:
        indexes.append(row.strip())

    File.close()

set_empty_bands()  # sostituisce il path delle bande non presenti nella directory con il path della banda vuota

# creiamo le immagini per ogni timestamp e le inseriamo in timeSeries
images = TimeSeries.TimeSeries()  # creo l'oggetto time series
for image in filePaths.keys():  # esegue tutto per ogni timestamp
    im = Image.Image()
    for band in filePaths[image].keys():  # esegue tutto per ogni banda
        temp = Band.Band(filePaths[image][band])  # crea la banda utilizzando il file nel relativo path
        im.set_band(temp)  # inserisce la banda nell'immagine
    images.add_image(im)


def create_indexes():
    """ Calcola gli indici e li inserisce all'interno delle rispettive immagini

    :returns: lista in cui in ogni posizione è presente un immagine contenente indici al posto delle bande
    """
    return images.createIndexTimeSeries(indexes)


def add_indexes_to_images(images, imagesIndexes):
    """ Crea un unico tensore contenente in ogni immagine sia le bande che gli indici

    :param images: contiene tutte le immagini che a loro volta contengono le bande
    :type images: TimeSeries
    :param imagesIndexes: contiene tutte le immagini che a loro volta contengono tutti gli indici calcolati su di essa
    :type imagesIndexes: TimeSeries
    :return: restituisce un unica time series contenente in ciascuna immagine sia le bande che gli indici
    :rtype: TimeSeries
    """
    dataset = copy.deepcopy(images)
    for indexes, image in zip(imagesIndexes.get_images(), dataset.get_images()):
        for index in indexes.get_tensor():
            image.set_band(index)

    return dataset


def split_test_train(percentualSplit, mask, dataset):
    """ Divide i pixel di ciascuna banda in train e test set

    :param percentualSplit: percentuale del training set rispetto l'intero dataset
    :type percentualSplit: Float
    :param mask: maschera contenente in posizione ZERO le posizioni dei pixel validi
    :type mask: Mask
    :param dataset: time series contenente in ciascuna immagine sia le bande che gli indici.
    :return: X_train contiene in ciascuna banda i pixel per il training set.
             X_test contiene in ciascuna banda i pixel per il test set.
             y_train contiene i risultati per i pixel del training set.
             y_test contiene i risultati per i pixel del test set.
    :rtype: (TimeSeries, TimeSeries, Label, Label)
    """
    maskTrain_TestSet = mask.split(percentualSplit)  #
    # come training set, in 1 quelle dei pixel per il test set
    dataset.set_label(label)
    X_train, y_train = dataset.get_dataset(
        maskTrain_TestSet[0])  # estraggo il training set tramite la relativa maschera
    X_test, y_test = dataset.get_dataset(maskTrain_TestSet[1])  # estraggo il test set tramite la relativa maschera

    return X_train, X_test, y_train, y_test


def calculate_mi(X_train, timeStamps):
    """ Calcola i valori MI sul training set e li inserisce in ordine decrescente all'interno un file Excel

    :param X_train: contiene in ciascuna banda i pixel per il training set
    :type X_train: TimeSeries
    :param timeStamps: lista contenente i nomi di ciascun times tamp
    :type timeStamps: List
    """
    MiResults = []  # in ogni posizione contiene la lista con tutti i valori MI calcolati su ciascuna banda e indice in un Time stamp
    for image in X_train.get_images():
        MiResults.append(image.MI(y_train))

    # per ogni banda calcoliamo i valori di MI derivanti dalla concatenazione di se stessa nei vari time stamp
    X_train.set_label(y_train)
    X_train.set_mask(mask)
    concatenationBands = X_train.MI()
    MiResults.append(concatenationBands)

    sheetsNames = timeStamps.copy()
    sheetsNames.append("concatenation")
    utility = Utility.Utility()
    utility.createExcels(bands, indexes, MiResults, sheetsNames)


def read_excel(fileName, timeStamps, X_train, X_test):
    """ Legge file Excel contenente i valori MI in ordine decrescente e ricrea train e test set in base al rank

    :param fileName: nome del file Excel
    :type fileName: String
    :param timeStamps: lista contenente i nomi di ciascun times tamp
    :type timeStamps: List
    :param X_train: contiene in ciascuna banda i pixel per il training set
    :type X_train: TimeSeries
    :param X_test: contiene in ciascuna banda i pixel per il test set
    :type X_train: TimeSeries
    :returns: ID è una lista in cui in ogni posizione sono presenti i nomi degli indici o bande selezionate per addestrare il classificatore.
              ranked_train lista con all'interno i dati del training set inseriti in ordine decrescente.
              ranked_test lista con all'interno i dati del test set inseriti in ordine decrescente.
    :rtype: (List, List, List)
    """
    # Leggo il file excel
    file_excel = openpyxl.load_workbook(fileName)
    ID = []  # lista in cui in ogni posizione sono presenti i nomi degli indici o bande selezionate per addestrare il classificatore
    for sheetName in timeStamps:
        sheet = file_excel[sheetName]
        i = 1
        bandsID = []
        while sheet.cell(row=i, column=1).value is not None:
            row = sheet.cell(row=i, column=1).value
            bandsID.append(row)
            i += 1
        ID.append(bandsID)

    # ricostruisco il training e test set in base al rank
    ranked_train = []
    ranked_test = []
    for image_train, image_test, bands_id in zip(X_train.get_images(), X_test.get_images(), ID):
        trainValues = []
        testValues = []
        for id in bands_id:
            i = 0
            for id_train_test in (bands + indexes):
                if id == id_train_test:
                    trainValues.append(image_train.get_band(index=i).get_band())
                    testValues.append(image_test.get_band(index=i).get_band())
                    break
                i += 1
        ranked_train.append(np.concatenate(trainValues, axis=1))
        ranked_test.append(np.concatenate(testValues, axis=1))

    return ID, ranked_train, ranked_test


imagesIndexes = create_indexes()

dataset = add_indexes_to_images(images, imagesIndexes)

label = Label.Label(labelPath)  # creo l'oggetto label
mask = Mask.Mask(label.get_flatten())
X_train, X_test, y_train, y_test = split_test_train(percentualSplit, mask, dataset)

calculate_mi(X_train, timeStamps)

idBand, rankedBands_train, rankedBands_test = read_excel('Mi_Bands.xlsx', timeStamps, X_train, X_test)
idIndexes, rankedIndexes_train, rankedIndexes_test = read_excel('Mi_Indexes.xlsx', timeStamps, X_train, X_test)
ID, ranked_train, ranked_test = read_excel('Mi_Values.xlsx', timeStamps, X_train, X_test)

utility = Utility.Utility()
utility.create_validation_excel()   # creo il file excel che conterrà i risultati dei test

models = ['RF', 'XGB']
for model in models:
    if model == 'RF':
        classifier = RandomForest.RandomForest()

    elif model == 'XGB':
        classifier = XGBoost.XGBoost()

    # N.B: ogni volta che in questa sezione si sceglie un opzione i risultati verranno stampati a video e salvati nel file Excel
    options = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j']
    cv = ClassifierValidator()
    for option in options:
        if option == 'a':
            # solo bande nell'ultimo time stamp
            classifier.fit(rankedBands_train[-1:][0], y_train.get_label())
            predictions = classifier.predict(rankedBands_test[-1:][0])

        elif option == 'b':
            # solo bande in tutti i time stamp
            classifier.fit(np.concatenate(rankedBands_train, axis=1), y_train.get_label())
            predictions = classifier.predict(np.concatenate(rankedBands_test, axis=1))

        elif option == 'c':
            # solo indici nell'ultimo time stamp
            classifier.fit(rankedIndexes_train[-1:][0], y_train.get_label())
            predictions = classifier.predict(rankedIndexes_test[-1:][0])

        elif option == 'd':
            # solo indici in tutti i time stamp
            classifier.fit(np.concatenate(rankedIndexes_train, axis=1), y_train.get_label())
            predictions = classifier.predict(np.concatenate(rankedIndexes_test, axis=1))

        elif option == 'e':
            # bande e indici nell'ultimo time stamp
            classifier.fit(ranked_train[-1:][0], y_train.get_label())
            predictions = classifier.predict(ranked_test[-1:][0])

        elif option == 'f':
            # bande e indici in tutti i time stamp
            classifier.fit(np.concatenate(ranked_train, axis=1), y_train.get_label())
            predictions = classifier.predict(np.concatenate(ranked_test, axis=1))

        elif option == 'g':
            # bande e top 10 indici su tutti i time stamp
            new_rankedIndexes_train = np.concatenate(([image[:, :10] for image in rankedIndexes_train]), axis=1)
            new_rankedBands_train = np.concatenate(rankedBands_train, axis=1)
            classifier.fit(np.concatenate((new_rankedBands_train, new_rankedIndexes_train), axis=1), y_train.get_label())

            new_rankedIndexes_test = np.concatenate(([image[:, :10] for image in rankedIndexes_test]), axis=1)
            new_rankedBands_test = np.concatenate(rankedBands_test, axis=1)
            predictions = classifier.predict(np.concatenate((new_rankedBands_test, new_rankedIndexes_test), axis=1))

        elif option == 'h':
            # bande e top 20 indici su tutti i time stamp
            new_rankedIndexes_train = np.concatenate(([image[:, :20] for image in rankedIndexes_train]), axis=1)
            new_rankedBands_train = np.concatenate(rankedBands_train, axis=1)
            classifier.fit(np.concatenate((new_rankedBands_train, new_rankedIndexes_train), axis=1), y_train.get_label())

            new_rankedIndexes_test = np.concatenate(([image[:, :20] for image in rankedIndexes_test]), axis=1)
            new_rankedBands_test = np.concatenate(rankedBands_test, axis=1)
            predictions = classifier.predict(np.concatenate((new_rankedBands_test, new_rankedIndexes_test), axis=1))

        elif option == 'i':
            # top 20 bande e indici sull'ultimo time stamp
            classifier.fit(ranked_train[-1][:, :20], y_train.get_label())
            predictions = classifier.predict(ranked_test[-1][:, :20])

        elif option == 'j':
            # top 20 bande e indici su ogni time stamp
            new_ranked_train = np.concatenate(([image[:, :20] for image in ranked_train]), axis=1)
            classifier.fit(new_ranked_train, y_train.get_label())

            new_ranked_test = np.concatenate(([image[:, :20] for image in ranked_test]), axis=1)
            predictions = classifier.predict(new_ranked_test)

        validation = cv.validation(y_test.get_label(), predictions)
        ConfusionMatrix = cv.confusion_matrix(y_test.get_label(), predictions)

        utility.add_validation(validation, model + " -" + option)

        print('\n\n\nValidation:')
        for key, value in validation.items():
            print(key, value)

        print("\nconfusion matrix:\n", ConfusionMatrix)