from openpyxl import Workbook
from openpyxl import load_workbook


class Utility:

    def __init__(self):
        pass

    def createExcels(self, idBands, idIndexes, results, sheetsNames):
        wb = Workbook()  # creo file excel
        wb.save('Mi_Bands.xlsx')
        wb.save('Mi_Indexes.xlsx')
        wb.save('Mi_Values.xlsx')

        # creo un excel con il rank dei valori MI delle bande
        wb = load_workbook('Mi_Bands.xlsx')
        for sheetName, imageResult in zip(sheetsNames, results):
            sheet = wb.create_sheet(sheetName)  # creo il foglio col nome del timestamp
            i = 1

            bandsResult = imageResult[:len(idBands)]  # prende solo i valori MI delle bande
            # effettuiamo il rank dei valori MI delle bande
            zipped_lists = zip(bandsResult, idBands)
            sorted_zipped_lists = sorted(zipped_lists, reverse=True)
            sortedResult, sortedID = zip(*sorted_zipped_lists)

            # inseriamo i valori nel file
            for id, value in zip(sortedID, sortedResult):
                if value != 0:  # salto le bande inutilizzate
                    sheet.cell(row=i, column=1).value = id
                    sheet.cell(row=i, column=2).value = value
                    i += 1
        wb.remove(wb.worksheets[0])  # elimino il primo sheet che viene generato automaticamente
        wb.save('Mi_Bands.xlsx')

        # creo un excel con il rank dei valori MI degli indici
        wb = load_workbook('Mi_Indexes.xlsx')
        for sheetName, imageResult in zip(sheetsNames, results):
            sheet = wb.create_sheet(sheetName)  # creo il foglio col nome del timestamp
            i = 1

            indexesResult = imageResult[len(idBands):]  # prende solo i valori MI degli indici
            # effettuiamo il rank dei valori MI degli indici
            zipped_lists = zip(indexesResult, idIndexes)
            sorted_zipped_lists = sorted(zipped_lists, reverse=True)
            sortedResult, sortedID = zip(*sorted_zipped_lists)

            # inseriamo i valori nel file
            for id, value in zip(sortedID, sortedResult):
                if value != 0:  # salto le bande inutilizzate
                    sheet.cell(row=i, column=1).value = id
                    sheet.cell(row=i, column=2).value = value
                    i += 1
        wb.remove(wb.worksheets[0])  # elimino il primo sheet che viene generato automaticamente
        wb.save('Mi_Indexes.xlsx')

        # creo excel con il rank di tutti i valori MI
        wb = load_workbook('Mi_Values.xlsx')
        for sheetName, imageResult in zip(sheetsNames, results):
            sheet = wb.create_sheet(sheetName)  # creo il foglio col nome del timestamp
            i = 1

            # effettuiamo il rank dei valori MI
            zipped_lists = zip(imageResult, idBands + idIndexes)
            sorted_zipped_lists = sorted(zipped_lists, reverse=True)
            sortedResult, sortedID = zip(*sorted_zipped_lists)

            # inseriamo i valori nel file
            for id, value in zip(sortedID, sortedResult):
                if value != 0:  # salto le bande inutilizzate
                    sheet.cell(row=i, column=1).value = id
                    sheet.cell(row=i, column=2).value = value
                    i += 1
        wb.remove(wb.worksheets[0])  # elimino il primo sheet che viene generato automaticamente
        wb.save('Mi_Values.xlsx')

    def create_validation_excel(self):
        wb = Workbook()
        sheet = wb.active

        # inseriamo i nomi delle metriche nella prima riga del file
        parameters = ['precision class1', 'recall class1', 'f1-score class1', 'support class1', 'precision class2', 'recall class2', 'f1-score class2', 'support class2', 'precision class3', 'recall class3', 'f1-score class3', 'support class3', 'precision class4', 'recall class4', 'f1-score class4', 'support class4', 'accuracy', 'macro av. F1-score', 'weigthed avg F1-score', 'weigthed avg recall']
        i = 2
        for parameter in parameters:
            sheet.cell(row=1, column=i).value = parameter
            i += 1

        wb.save('validation.xlsx')

    def add_validation(self, validation, name):
        wb = load_workbook('validation.xlsx')
        sheet = wb.active

        rowIndex = 1
        while 1 != 0:
            if sheet[rowIndex][1].value is None:   # scorro le righe fino a trovare quella vuota
                break
            rowIndex += 1

        # inseriamo i risultati delle metriche nelle relative colonne
        i = 1
        sheet.cell(row=rowIndex, column=i).value = name  # inserisco nella prima colonna il nome dell'operazione
        i += 1
        for dictionary in validation:
            if type(validation[dictionary]) == dict:
                for key in validation[dictionary]:
                    sheet.cell(row=rowIndex, column=i).value = validation[dictionary][key]
                    i += 1
            else:
                sheet.cell(row=rowIndex, column=i).value = validation[dictionary]
                i += 1
                break

        sheet.cell(row=rowIndex, column=i).value = validation['macro avg']['f1-score']
        i += 1
        sheet.cell(row=rowIndex, column=i).value = validation['weighted avg']['f1-score']
        i += 1
        sheet.cell(row=rowIndex, column=i).value = validation['weighted avg']['recall']

        wb.save('validation.xlsx')